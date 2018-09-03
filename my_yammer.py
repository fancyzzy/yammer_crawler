#!/usr/bin/env python2
#!encoding: utf-8

import sys
from datetime import datetime
import openpyxl
import os

import my_database
import my_crawler

#solve flask multithread trigger
#import pythoncom
#pythoncom.CoInitialize()


class My_Yammer():

    def __init__(self, access_token):

        self.my_db = my_database.My_Database()
        self.my_crawler = my_crawler.My_Crawler(access_token)
        print("my_yammer init finished")

    ##########__init__()################################


    def pull_all_messages(self, group_id, interval=1):

        print("start pull_all_messages, group_id = {}".format(group_id))
        all_messages = self.my_crawler.download_all_messages(group_id, interval)

        #save to db
        if all_messages != None:
            self.my_db.save_group_messages(all_messages, str(group_id))
            print("Messages data saved successfully.")
            return True
        else:
            print("No messages data saved.")
            return False

    ############pull_all_messages()###############################


    def pull_newer_messages(self, group_id, interval=1):

        print("start pull_newer_messages, group_id = {}".format(group_id))
        existed_messages = self.my_db.get_group_messages(str(group_id))

        #Continue to download messages that are newer that the latest existed message
        newer_messages = None
        if existed_messages != None:
            newer_than_id = existed_messages["messages"][0]["id"]
            newer_messages = self.my_crawler.download_newer_messages(group_id, newer_than_id, interval)

            #save to db
            if newer_messages != None:
                #merge newer_message to existed_messages
                self.my_db.update_group_messages(existed_messages, newer_messages, group_id)
                print("{} New messages data updateded successfully.".format(len(newer_messages)))
                return True
            else:
                print("No messages data updateded.")
                return False
        #No existed messages, start to download for all
        else:
            print("No existed data, pull all the messages")
            self.pull_all_messages(group_id, interval)
    #################pull_newer_messages()#########################


    def pull_all_users(self, group_id, interval=1):
        '''
        download all the user general info

        :param group_id:
        :param interval:
        :return:
        '''
        dict_users = self.my_crawler.download_all_users(group_id, interval)

        #save to db
        if dict_users != None:
            self.my_db.save_group_users(dict_users, str(group_id))
            return True
        else:
            return False
    ############pull_all_users()###############################

    def pull_all_users_details(self, group_id, interval=1):
        '''
        download all the users detailed info and save each one into a json file

        :param group_id:
        :param interval:
        :return:
        '''
        print("Start to download each user detailed info of group {}".format(group_id))
        existed_users = self.my_db.get_group_users(str(group_id))

        #download each user's detailed info
        n = 0
        for user_dict in existed_users["users"]:

            print("Download batch: {}".format(n+1))
            dict_user = self.my_crawler.download_user_details(user_dict)
            self.my_db.save_group_user_details(dict_user, str(group_id))
            n += 1
        return True
    ###############pull_all_users_details()####################


    def pull_all_users_and_details(self, group_id, interval=1):
        '''
        download all users general info json and each user details json

        :param group_id:
        :param interval:
        :return:
        '''

        self.pull_all_users(group_id, interval)
        self.pull_all_users_details(group_id, interval)

    ################pull_all_users_and_details()################


    def get_group_name(self, group_id):

        existed_messages = self.my_db.get_group_messages(str(group_id))
        if existed_messages == None:
            print("Group data is not existed yet")
        else:
            return existed_messages["meta"]["feed_name"]
    ########get_group_name###########################################


    def get_group_messages(self, group_id):
        '''
        :param group_id:
        :return: existed_messages like https://www.yammer.com/api/v1/messages/in_group/15273590.json
        '''

        existed_messages = self.my_db.get_group_messages(str(group_id))
        #logic, algorithm

        return existed_messages
    #############get_group_message()###############################


    def get_group_users(self, group_id):
        '''
        existed_users like https://www.yammer.com/api/v1/users/in_group/15273590.json
        '''

        existed_users = self.my_db.get_group_users(str(group_id))
        print("DEBUG existed_users: {}".format(existed_users))

        #convert to id:user_data dict
        users_info = {}
        for user in existed_users["users"]:
            users_info[user["id"]] = user

        return users_info
    ########get_group_users()#####################################


    def get_user_info(self, user_id, group_id=None):
        '''
        Get one user general information

        :param user_id:
        :param group_id:
        :return:
        '''

        return self.my_db.get_user_info(user_id, str(group_id))

    ##############get_user_info()##################################


    def get_groups(self):
        return self.my_db.get_all_groups_name_id()
    ##################get_current_group_names()####################


    #Game
    def get_group_rank(self, group_id, letter_num=1, least_comment_num=1, end_date=None, start_date=None,\
                       is_sorted_for_post=False):
        '''
        Get a sorted list which contatin user name, message num which is the key to rank
        Only return those users who had sent messages. For those not sent even one message,
        The return list does not contatin them

        :param group_id:
        :param letter_num: letter number of a message content
        :param least_comment_num: number of messages sent
        :param end_date:   liek '2018/08/07'
        :param start_date: '2018/02/01'
        :return: list: [[id, name, comment, post],...]
        '''
        print("Start show group rank with at least letter_num: {}, least_comment_num: {}, from date: {} to {}".\
              format(letter_num, least_comment_num, end_date, start_date))

        #{id:[total_message, post_message],...}
        d_users = {}
        #result_list = [[id,total_message_number, post_message_number],...]
        result_list = []
        n = 0
        n_post = 0

        messages = self.get_group_messages(group_id)

        if messages == None:
            print("nothing")
            return

        for message in messages["messages"]:

            created_date = message["created_at"].split()[0]

            if (start_date !=None) and (created_date < start_date):
                break
            if (end_date != None) and (created_date > end_date):
                continue

            sender_id = message["sender_id"]
            content = message["body"]["plain"]
            is_replied = message["replied_to_id"]
            #message_type = message["message_type"]
            #print("DEBUG id: {}, sender_id: {}, is_replied: {}".format(message["id"], sender_id, is_replied))

            if len(content.split()) >= letter_num:
                if sender_id not in d_users.keys():
                    d_users.setdefault(sender_id,[0,0])

                #is a post
                if is_replied == None:
                    d_users[sender_id][1] += 1
                    n_post += 1
                #is a comment/update
                else:
                    d_users[sender_id][0] += 1
                    n += 1

        #[[id, comments, post], ...]
        result_list = [[x,d_users[x][0],d_users[x][1]] for x in d_users.keys() if d_users[x][0] >= least_comment_num]

        if is_sorted_for_post:
            #for post
            idx = 2
        else:
            #for comment
            idx = 1
        ranked_list = sorted(result_list, key=lambda x:x[idx], reverse=True)

        #get user name by id and append user id and  photo
        user_info = self.get_group_users(group_id)
        unknown_num = 0
        for user in ranked_list:
            user_id = user[0]
            if user_id in user_info.keys():
                user_name = user_info[user_id]["full_name"]
                #Simple name
                user_name = user_name.split(', ')[0].upper() +' '+ user_name.split(', ')[1].split(' ')[0]
                user[0] = user_name
                user_photo = user_info[user_id]["mugshot_url"]
                user.append(user_photo)
            else:
                print("warning, unknown user: {} detected".format(user_id))
                unknown_url = 'https://www.yammer.com/api/v1/users/' + str(user_id) + '.json'
                print("check {} to find this user".format(unknown_url))
                user[0] = 'unknown user'
                none_photo = "https://mug0.assets-yammer.com/mugshot/images/48x48/no_photo.png"
                user.append(none_photo)
                unknown_num += 1

            #insert id for future index purpose
            user.insert(0, user_id)

        if start_date == None:
            start_date = 'the ever beginning.'
        group_name = self.get_group_name(group_id)
        print("In the group '{}',".format(group_name))
        print("Totally {} comments and  {} posts from date {} back to {}".format(n, n_post, end_date, start_date))
        print("Where {} people sent at least {} comments".format(len(result_list), least_comment_num))
        if unknown_num > 0:
            print("%d unknown user."%(unknown_num))

        #rank_field = ["Id", "Full_Name", "Updats", "Posts", "Photo"]

        return ranked_list
    #############get_group_rank()##################################################


    def export_users_details_to_excel(self,group_id):
        '''
        export users details to an easy readable xlsx file

        :param group_id:
        :return:
        '''
        print("Export to users details to excel of group: {}".format(group_id))
        dict_list = self.my_db.get_users_details(str(group_id))

        dest_folder = my_database.DATA_PATH

        total_num = len(dict_list)

        if total_num == 0:
            print("No user details found for group {}".format(group_id))
            return False

        else:
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = "Users"
            column_fields = ['Id', 'State', 'Full_Name', 'Job_Title', 'Department', \
                             'Email', 'Phone', 'Photo', 'Interests', 'Expertise', 'Stats']

            field_num = len(column_fields)
            start_row = 5
            start_col = 2

            #Write the first header
            for j in range(field_num):
                sheet.cell(row=start_row, column=start_col+j).value = column_fields[j]

            #Write data
            for i in range(total_num):
                # user_dict format see:
                # https://www.yammer.com/api/v1/users/1568779836.json
                user_d = dict_list[i]

                for j in range(field_num):
                    field = column_fields[j].lower()
                    if field == "phone":
                        val = str(user_d["contact"]["phone_numbers"])
                    elif field == "photo":
                        val = str(user_d["mugshot_url"])
                    else:
                        val = str(user_d[field])
                    sheet.cell(row=start_row+1+i, column=start_col+j).value = val

            excel_name = 'group_%s_users.xlsx'%(group_id)
            folder_path = os.path.join(my_database.DATA_PATH, 'group_%s'%(group_id))
            excel_path =os.path.join(folder_path, excel_name)
            wb.save(filename=excel_path)
            print("Export finished at {}".format(excel_path))
            return True
    ################export_users_details_to_excel()########################################


if __name__ == '__main__':

    print("start my_yammer")

    email_add = ''
    csl = ''
    pwd = ''
    group_id = 15273590
    #group_id = '12562314' #Qingdao

    access_token = '592-FnmLDb1cF0zMgyj32jnz0w'
    my_yammer = My_Yammer(access_token)

    my_yammer.pull_newer_messages(group_id, interval=1)
    my_yammer.pull_all_users(group_id, interval=1)

    #str_now = datetime.now().strftime("%Y/%m/%d")
    #my_yammer.get_group_rank(group_id, letter_num=0, end_date=str_now, start_date=None)

    #my_yammer.pull_all_users_details(group_id, interval=1)

    #group_id = '12562314' #Qingdao
    #my_yammer.pull_all_users_and_details(group_id, interval=1)
    #my_yammer.pull_newer_messages(group_id, interval=1)
    str_now = datetime.now().strftime("%Y/%m/%d")
    ranked_list = my_yammer.get_group_rank(group_id, letter_num=0, least_comment_num=1, end_date=str_now, \
                                           start_date=None, is_sorted_for_post=True)

    #my_yammer.export_users_details_to_excel(group_id)
    if ranked_list != None:
        for item in ranked_list:
            print(item)

    print("done")